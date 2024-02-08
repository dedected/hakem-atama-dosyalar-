import pandas as pd
import numpy as np
import re
import random
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

'''
def hakem_atamasi():
    # Hakem şartları dosyalarını açın
    epak_listesi = "EPAK listesi.xlsx"
    hakem_mazeret_listesi = "hakem mazeret listesi.xlsx"
    antrenmana_katılanlar_listesi = "antrenmana katılanlar listesi.xlsx"
    antrenman_günü_maçı_olanlar_listesi = "antrenman günü maçı olanlar listesi.xlsx"

    # Hakem şartları dosyalarını oku
    epak_listesi_oku = pd.read_excel(epak_listesi)
    hakem_mazeret_listesi_oku = pd.read_excel(hakem_mazeret_listesi)
    antrenmana_katılanlar_listesi_oku = pd.read_excel(antrenmana_katılanlar_listesi)
    antrenman_günü_maçı_olanlar_listesi_oku = pd.read_excel(antrenman_günü_maçı_olanlar_listesi)

    # Maç atanacak hakemler listesi oluştur
    maç_atanacak_hakemler = []

    # EPAK'a katılıp katılmadığını kontrol ederek katılanların bir listesini oluştur
    epak_katılan_hakemler = []
    for index, satir in epak_listesi_oku.iterrows():
        if isinstance(satir.iloc[5], (str,float)) and str(satir.iloc[5]).upper() == "X":
            epak_katılan_hakemler.append(satir.iloc[1])

    # Mazeret kontrol edilecek tarihi al
    tarih = input("Mazeret kontrol edilecek tarihi girin(gün.ay.yıl olarak): ")

    # girilen tarihte hakemin mazereti var mı diye kontrol etmek için
    tarihler = hakem_mazeret_listesi_oku.iloc[1::4, 2].tolist() # dosyanın 3. sütununda 3.satırdan itibaren her 4 satırda bir mazeret tarihleri yazıyor
    isimler = hakem_mazeret_listesi_oku.iloc[1::4, 0].tolist() # dosyanın 1. sütununda 3.satırdan itibaren her 4 satırda bir isimler yazıyor

    # girilen tarihte mazereti olan hakemleri listeden çıkarmak için silinecek hakemler adlı geçici bir liste açıp girilen tarihte mazereti olan hakemleri bu listeye koyma
    silinecek_hakemler = []
    for hakem, mazeret_tarihi in zip(isimler, tarihler):
                if mazeret_tarihi == tarih:
                    silinecek_hakemler.append(hakem)

    # antrenman günü maça çıkanları al ve tekrarlanan isimleri tek bir değere düşür
    antrenman_günü_maçı_olanlar = []
    antrenman_günü_maçı_olanlar_isimleri = antrenman_günü_maçı_olanlar_listesi_oku.iloc[2:, [7, 8, 9]].values.tolist()

    for hakemler in antrenman_günü_maçı_olanlar_isimleri:
        for hakem in hakemler:
            if isinstance(hakem ,str):
                antrenman_günü_maçı_olanlar.append(hakem)

    unique_set = set(antrenman_günü_maçı_olanlar)
    antrenman_günü_maçı_olanlar_unique = list(unique_set)

    # antrenmana katılanları al
    antrenmana_katılanlar = antrenmana_katılanlar_listesi_oku.iloc[:, 0].tolist()

    # bütün şartları kontrol ederek girilen tarihte maç atanması uygun olan hakemleri listele
    for hakem in epak_katılan_hakemler:
        if hakem in silinecek_hakemler:
            pass
        elif hakem not in silinecek_hakemler:
            if hakem in antrenman_günü_maçı_olanlar_unique or hakem in antrenmana_katılanlar:
                maç_atanacak_hakemler.append(hakem)

    # maç atanacak hakemler listesini excele dönüştür
    df = pd.DataFrame({"Hakemler": maç_atanacak_hakemler})
    df.to_excel("maç atanacak hakemler.xlsx", index=False, header=False)

    return maç_atanacak_hakemler, silinecek_hakemler


def mac_atamasi():

    # önceki fonksiyondan maç atanmaya uygun hakemleri hakem_listesi listesine ekle
    hakem_listesi = []  
    hakem_listesi, _ = hakem_atamasi()  # Hakem listesini baştan al

    # excel dosyalarını aç ve oku
    mac_listesi_exceli = "maçlar.xlsx"
    mac_listesi_exceli_oku = pd.read_excel(mac_listesi_exceli)
    mac_listesi_secilen_sutunlar = mac_listesi_exceli_oku.iloc[25:, [2, 3, 4, 5, 6]]

    # hakem listelerini oluştur
    atanacak_hakemler = []
    hakemli_mac_listesi = []
    kalan_hakemler = []  # Atanamayan (kalan) hakemlerin listesi

    # maç atanacak tarihi al
    tarih2 = str(input("Maç atanacak tarihi girin (örn: 10 ARALIK 2023 (seçilen tarihi bu formatta giriniz)): "))

    # maçlar excelinde maç atanacak günün maçlarını almak için oluşturulan regex araması
    tarih2_yil_ay = tarih2.split(" ")[2:]
    tarih2_regex1 = r"\b" + re.escape(tarih2) + r"\b"
    regex_pattern1 = re.compile(tarih2_regex1)

    tarih2_regex2 = r"\b" + re.escape(" ".join(tarih2_yil_ay)) + r"\b"
    regex_pattern2 = re.compile(tarih2_regex2)

    matching_rows_indices = []

    for index, row in mac_listesi_exceli_oku.iterrows():
        value_in_second_column = str(row.iloc[1])
        value_in_third_column = str(row.iloc[2])

        match_result_second_column = regex_pattern1.search(value_in_second_column)
        if match_result_second_column:
            matching_rows_indices.append(index + 2)
        else:
            match_result_third_column = regex_pattern1.search(value_in_third_column)
            if match_result_third_column:
                matching_rows_indices.append(index + 2)

    print("Matching rows indices:")
    print(matching_rows_indices)

    if matching_rows_indices and matching_rows_indices[-1] < len(mac_listesi_exceli_oku) - 1:
        next_index_to_search = matching_rows_indices[-1] + 2
        print(next_index_to_search)

        for index, row in mac_listesi_exceli_oku.iloc[next_index_to_search:].iterrows():
            value_in_second_column = str(row.iloc[1])
            value_in_third_column = str(row.iloc[2])

            match_result_second_column = regex_pattern2.search(value_in_second_column)
            if match_result_second_column:
                matching_rows_indices.append(index + 2)
                break
            else:
                match_result_third_column = regex_pattern2.search(value_in_third_column)
                if match_result_third_column:
                    matching_rows_indices.append(index + 2)
                    break

            continue

    print("Matching rows indices after additional search:")
    print(matching_rows_indices)

    # günün başlangıç satır sayısını ve bitiş satır sayısını al. eğer son tarih seçildiyse dosyanın son satır sayısını bitiş satırı olarak al
    day_start_line = matching_rows_indices[0]
    if len(matching_rows_indices) > 1:
        day_end_line = matching_rows_indices[1]
    else:
        day_end_line = len(mac_listesi_exceli_oku) + 2

    print(day_start_line)
    print(day_end_line)


        #  MAÇ ATAMA
    # gerekli dosyaları aç ve oku
    mac_listesi = mac_listesi_secilen_sutunlar.values.tolist()
    mac_listesi_secilen_sutunlar = mac_listesi_exceli_oku.iloc[day_start_line - 1:day_end_line -2, [2, 3, 4, 5, 6]]
    mac_listesi_secilen_sutunlar.columns = ['Saha', 'Saat', 'Klasman', 'İç Saha', 'Deplasman']  # İlgili sütun isimlerini güncelle
    print(mac_listesi_secilen_sutunlar)


    # maç_atanacak_hakemler listesini oluşturduktan sonra rastgele karıştır:
    random.shuffle(hakem_listesi)

        # maç atama döngüsü
    # aynı sahadaki maçlara aynı hakemin/hakemlerin atanabilmesi için saha takibi için previous_field değişkeni
    previous_field = None
    # u13 maçlarında aynı sahada 2'den fazla maç varsa her  maçta bir atanan hakemin değişmesi için u13_sayac değişkeni
    u13_sayac = 0
    # u13 olmayan maçlarda aynı sahada 1'den fazla maç varsa atanan 3 hakemin diğer maçta sıralama değiştirmesi için oluşturulan u13_olmayan_sayac değişkeni
    u13_olmayan_sayac = 0
    # atamada u13 maçlarından u13 olmayan maçlara geçişte hakemlerde karışıklık olmaması için önceki klasmanı takip eden previous_klasman değişkeni
    previous_klasman = None

    for index, row in mac_listesi_secilen_sutunlar.iterrows():
        maclar = row
        klasman = row['Klasman'] 
        current_field = row['Saha']
        
        print("saha1:",current_field)
        print("saha2:",previous_field)

        if 'U 13' in klasman:
            # saha önceki saha ile aynıysa aynı hakemi kullan
            if current_field == previous_field and hakem_listesi:
                # aynı sahadaki u13 maçlarında aynı hakem atanan maç sayısı 2'den azsa aynı hakemi ekle
                if u13_sayac < 1:
                    hakem = atanacak_hakemler[-1]
                    u13_sayac += 1
                # aynı sahadaki u13 maçlarında aynı hakem atanan maç sayısı 2'den fazlaysa aynı hakemi ekle    
                elif u13_sayac >= 1:
                    hakem = hakem_listesi.pop(0)
                    atanacak_hakemler.append(hakem)
                    u13_sayac = 0

            # saha önceki ile aynı değilse u13_sayac'ı sıfırla ve atamak için hakem listesinden yeni hakem al        
            else:
                u13_sayac = 0
                if hakem_listesi:
                    hakem = hakem_listesi.pop(0)
                    atanacak_hakemler.append(hakem)
                else:
                    # Handle the case where hakem_listesi is empty
                    hakem = None

            # Append the entire row along with the assigned referee to hakemli_mac_listesi
            maclar_with_hakem = row.to_dict()
            maclar_with_hakem['Hakem'] = hakem
            hakemli_mac_listesi.append(maclar_with_hakem)

            # Update the previous field for the next iteration
            previous_field = current_field
            previous_klasman = klasman

        else:
            if current_field == previous_field and hakem_listesi:
                if 'U 13' in previous_klasman:
                # u13 olmayan maçlarda önceki klasman u13 ise atanacak hakemlere hakem listesinden yeni hakem al
                    if len(hakem_listesi) >= 3:
                        hakem1 = hakem_listesi.pop(0)
                        hakem2 = hakem_listesi.pop(0)
                        hakem3 = hakem_listesi.pop(0)
                        atanacak_hakemler.extend([hakem1, hakem2, hakem3])
                    else:
                        # Handle the case where hakem_listesi does not have enough elements
                        hakem1, hakem2, hakem3 = None, None, None
                elif 'U 13' not in previous_klasman:
                # u13 olmayan maçlarda önceki klasman u13 değil ise aynı hakemleri kullan
                    # hakemlerin aynı sahadaki diğer maçlarda yer değiştirme işlemleri
                    if len(hakem_listesi) >= 3:
                        if u13_olmayan_sayac < 1:
                            hakem1 = atanacak_hakemler[-2]
                            hakem2 = atanacak_hakemler[-1]
                            hakem3 = atanacak_hakemler[-3]
                            u13_olmayan_sayac += 1
                        elif u13_olmayan_sayac >= 1:
                            hakem1 = atanacak_hakemler[-1]
                            hakem2 = atanacak_hakemler[-3]
                            hakem3 = atanacak_hakemler[-2]
                            u13_olmayan_sayac = 0
                    else:
                        # Handle the case where hakem_listesi does not have enough elements
                        hakem1, hakem2, hakem3 = None, None, None
            else:
            # u13 olmayan maçlarda önceki klasman u13 değilse ve önceki saha mevcut saha ile aynı değilse atama işlemi
                if len(hakem_listesi) >= 3:
                    # hakem listesinin eleman sayısı 3'ten büyük veya 3 ise yeni 3 tane hakem ata
                    hakem1 = hakem_listesi.pop(0)
                    hakem2 = hakem_listesi.pop(0)
                    hakem3 = hakem_listesi.pop(0)
                    atanacak_hakemler.extend([hakem1, hakem2, hakem3])
                else:
                    # Handle the case where hakem_listesi does not have enough elements
                    hakem1, hakem2, hakem3 = None, None, None

            # Append the entire row along with the assigned referee to hakemli_mac_listesi
            maclar_with_hakem = row.to_dict()
            maclar_with_hakem['Hakem 1'] = hakem1
            maclar_with_hakem['Hakem 2'] = hakem2
            maclar_with_hakem['Hakem 3'] = hakem3
            hakemli_mac_listesi.append(maclar_with_hakem)

            # Update the previous field for the next iteration
            previous_field = current_field
            previous_klasman = klasman

    # Atanamayan (kalan) hakemleri ekle
    kalan_hakemler.extend(hakem_listesi)

    # Hakemli maçları yazdır
    print("Hakemli maçlar:")
    #print(hakemli_mac_listesi)
    for maclar in hakemli_mac_listesi:
        print(maclar)

    # Atanamayan (kalan) hakemleri yazdır
    print("Kalan Hakemler:")
    for hakem in kalan_hakemler:
        print(hakem)

    # Hakemli maçları içeren DataFrame oluşturun
    hakemli_mac_dataframe = pd.DataFrame(hakemli_mac_listesi)

    # Excel dosyasını yazdırırken bir ExcelWriter nesnesi kullanın
    with pd.ExcelWriter("hakemli_maclar.xlsx", engine="xlsxwriter") as writer:
        # DataFrame'i Excel dosyasına yazdırın, başlangıç satırını belirtin
        hakemli_mac_dataframe.to_excel(writer, index=False, sheet_name="Hakemli_Maclar", startrow=1)

        # Excel dosyasındaki yazılacak sayfa nesnesini alın
        worksheet = writer.sheets["Hakemli_Maclar"]

        # Tarih bilgisini yazmak için bir yazım nesnesi oluşturun
        date_format = writer.book.add_format({'num_format': 'dd mm yyyy'})

        # Başlık hücrelerini özelleştirme
        header_format = writer.book.add_format({'bold': True, 'font_size': 18, 'align': 'center', 'valign': 'vcenter'})
        worksheet.merge_range('A1:B1', 'Tarih: ' + tarih2, header_format)


    # Kalan hakemleri bir Excel dosyasına yazdır
    df_kalan_hakemler = pd.DataFrame({"Kalan Hakemler": kalan_hakemler})
    df_kalan_hakemler.to_excel("kalan_hakemler.xlsx", index=False, header=False)
'''
    
def mazeretsiz_hakemler():
    # Hakem şartları dosyalarını açın
    refrees_full_list = "hakemler tam liste.xlsx"
    yeni_mazeret_listesi = "yeni mazeret listesi.xlsx"

    # Hakem şartları dosyalarını oku
    refrees_full_list_read = pd.read_excel(refrees_full_list)
    yeni_mazeret_listesi_oku = pd.read_excel(yeni_mazeret_listesi)

    # hakem listesinden hakem adları ve klasmanlarını al
    refree_names = refrees_full_list_read.iloc[1:, 1]
    #refree_names = pd.Series(refree_names)
    refree_classification = refrees_full_list_read.iloc[1:, 2]
    refree_classification = pd.Series(refree_classification)

    # Mazeret kontrol edilecek tarihi al
    tarih3 = str(input("2-Mazeret kontrol edilecek tarihi girin(gün Ay yıl olarak): "))

    # girilen tarihte hakemin mazereti var mı diye kontrol etmek için
    tarihler = yeni_mazeret_listesi_oku.iloc[1:, 2].apply(lambda x: ' '.join(x.split()[:3])).tolist()
    isimler = yeni_mazeret_listesi_oku.iloc[1:, 1].tolist()

    # girilen tarihte mazereti olan hakemleri listeden çıkarmak için silinecek hakemler adlı geçici bir liste açıp girilen tarihte mazereti olan hakemleri bu listeye koy
    mazeretli_hakemler_list = []
    for hakem, mazeret_tarihi in zip(isimler, tarihler):
                if mazeret_tarihi == tarih3:
                    mazeretli_hakemler_list.append(hakem)
   
    # Yeni Excel dosyasını oluştur
    with pd.ExcelWriter('mazeretsiz hakemler.xlsx', engine='openpyxl') as writer:
        # DataFrame'i yaz
        df = pd.DataFrame({'AD SOYAD': refree_names, 'KLASMAN': refree_classification})
        df.to_excel(writer, sheet_name='Sheet1', index=False)

        # Mazeretli hakemlerin satırlarını renklendir
        workbook  = writer.book
        worksheet = workbook['Sheet1']

        # Mazeretli hakemlerin satırlarını renklendir
        index = 0
        for idx, refree in enumerate(refree_names):
            is_row_filled = False
            for col in range(1, worksheet.max_column + 1):
                fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                cell = worksheet.cell(row=idx + 2, column=col)
                if cell.fill.fill_type == "solid" and cell.fill.start_color.rgb == "FFFF00" and cell.fill.end_color.rgb == "FFFF00":
                    is_row_filled = True
                    break
    
            if is_row_filled:
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=idx + 2, column=col)
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.fill = fill
            elif refree in mazeretli_hakemler_list:
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=idx + 2, column=col)
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    cell.fill = fill
            index += 1
            print(f"satır = {idx} {is_row_filled}")
                    
                    



'''
secim = input("Maç ataması yapmak istiyor musunuz? Evet için e hayır için h yazınız: ").lower()

if secim == 'e':
    mac_atamasi()
    mazeretsiz_hakemler()
elif secim == 'h':
    mazeretsiz_hakemler()
else:
    print("Geçersiz karakter girdiniz!")
'''

mazeretsiz_hakemler()


